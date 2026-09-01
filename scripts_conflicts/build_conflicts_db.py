#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Construiește baza de date a conflictelor militare de pe teritoriul României
(sec. VIII î.Hr. – 1944) din modulele conflicts_part*.py și exportă:
  - CSV (UTF-8 cu BOM, importabil în Excel/Google Sheets)
  - JSON (pentru aplicații web)
  - XLSX (cu cel puțin 2 foi: date + statistici + dicționar de câmpuri)
  - RAPORT_STRUCTURARE.md

Sursa de bază: batalii.csv (fișierul utilizatorului).
Rulare:  python3 scripts_conflicts/build_conflicts_db.py
"""
import csv
import io
import json
import os
import statistics
import sys
from collections import Counter, defaultdict

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(SCRIPT_DIR)
OUT_DIR = os.path.join(ROOT, "data", "conflicte_militare")

sys.path.insert(0, SCRIPT_DIR)
from conflicts_part1_antiquity import EVENTS as E1
from conflicts_part2_hellenistic import EVENTS as E2
from conflicts_part3_roman import EVENTS as E3
from conflicts_part4_migrations import EVENTS as E4
from conflicts_part5_medieval import EVENTS as E5
from conflicts_part6_modern import EVENTS as E6

PARTS = [("Sec. VIII–IV î.Hr. (antichitate)", E1),
         ("Sec. III–I î.Hr. (perioada elenistică)", E2),
         ("Sec. I–IV d.Hr. (Dacia romană)", E3),
         ("Sec. V–X d.Hr. (migrații)", E4),
         ("Sec. XI–XV d.Hr. (medieval)", E5),
         ("Sec. XVI–XX d.Hr. (modern/contemporan)", E6)]

FIELDS = [
    ("id", "ID", "Identificator unic al înregistrării"),
    ("titlu", "Titlu", "Numele bătăliei / conflictului"),
    ("tip", "Tip conflict", "Tipul evenimentului (bătălie, invazie, asediu, revoltă etc.)"),
    ("data_start", "Data start", "Datare (cu precizie: zi/lună sau an; c. = circa)"),
    ("data_end", "Data end", "Sfârșitul perioadei (pentru campanii/războaie)"),
    ("an_start", "An start", "An numeric (negativ = î.Hr.) — pentru sortare/filtrare"),
    ("an_end", "An end", "An numeric final (negativ = î.Hr.)"),
    ("secol", "Secol", "Secolul din sursă (ex.: sec. XV d.Hr.)"),
    ("secol_n", "Secol (număr)", "Număr: negativ = î.Hr., pozitiv = d.Hr. — pentru filtrare"),
    ("locatie", "Locație", "Localizare: localitate/județ sau zonă aproximativă"),
    ("zona_aprox", "Locație aproximativă", "1 = zonă aproximativă/necunoscută; 0 = localizare mai precisă"),
    ("judet", "Județ actual", "Județul actual (— dacă nu se poate stabili)"),
    ("regiune", "Regiune istorică", "Regiunea istorică/geografică"),
    ("lat", "Latitudine", "Coordonată aproximativă (WGS84)"),
    ("lng", "Longitudine", "Coordonată aproximativă (WGS84)"),
    ("participanti", "Participanți", "Beligeranți/forțe implicate (normalizat în română)"),
    ("descriere", "Descriere", "Context și desfășurare, pe baza sursei și a istoriografiei"),
    ("rezultat", "Rezultat", "Cine a câștigat / consecințe"),
    ("teritoriu", "Pe teritoriul actual al României", "da / partial / nu (nu = eveniment de context cu participare românească)"),
    ("observatii", "Observații", "Incertitudini de datare/locație și note metodologice"),
    ("sursa", "Sursă", "Referința din batalii.csv (secția/paragraful)"),
]

TERITORIU_LABEL = {
    "da": "Da",
    "partial": "Parțial (graniță/zona învecinată)",
    "nu": "Nu (context regional)",
}


def build_events():
    events = []
    for part_name, part in PARTS:
        for ev in part:
            e = dict(ev)
            e.setdefault("observatii", "—")
            e.setdefault("sursa", "batalii.csv")
            e.setdefault("zona_aprox", 0)
            e.setdefault("judet", "—")
            e.setdefault("lat", None)
            e.setdefault("lng", None)
            # validare de bază
            assert e["titlu"] and e["data_start"] and e["locatie"] and e["participanti"] and e["descriere"], e
            events.append(e)
    # sortare cronologică: perioadă, apoi titlu
    def sort_key(e):
        return (e["an_start"], e["an_end"] or e["an_start"], e["titlu"])
    events.sort(key=sort_key)
    for i, e in enumerate(events, 1):
        e["id"] = "RM-%03d" % i
    return events


def write_csv(events):
    os.makedirs(OUT_DIR, exist_ok=True)
    path = os.path.join(OUT_DIR, "conflicte_militare_romania.csv")
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[k for k, _, _ in FIELDS], extrasaction="ignore")
        w.writeheader()
        w.writerows(events)
    return path


def write_json(events):
    path = os.path.join(OUT_DIR, "conflicte_militare_romania.json")
    payload = {
        "meta": {
            "titlu": "Baza de date a conflictelor militare de pe teritoriul României",
            "descriere": "Conflicte militare de pe teritoriul actual al României, sec. VIII î.Hr. – 1944",
            "sursa_principala": "batalii.csv (fișierul utilizatorului) + completări din istoriografia consacrată",
            "generat": "2026-09-01",
            "numar_inregistrari": len(events),
            "campi": {k: d for k, d, _ in FIELDS},
        },
        "conflicte": events,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    return path


def write_xlsx(events):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Foaia 1: Date
    ws = wb.active
    ws.title = "Conflicte"
    headers = [h for _, h, _ in FIELDS]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="B03A2E")
        c.alignment = Alignment(vertical="top", wrap_text=True)
    for e in events:
        row = []
        for k, _, _ in FIELDS:
            v = e.get(k, "")
            if k == "teritoriu":
                v = TERITORIU_LABEL.get(v, v)
            row.append(v)
        ws.append(row)
    widths = [10, 46, 26, 14, 14, 10, 10, 20, 12, 46, 14, 16, 34, 11, 11, 52, 90, 58, 26, 60, 44]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    ws.row_dimensions[1].height = 30
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # ---- Foaia 2: Statistici
    ws2 = wb.create_sheet("Statistici")
    ws2.append(["Dimensiune", "Valoare"])
    ws2.append(["Total înregistrări", len(events)])
    ws2.append(["Pe teritoriul actual al României", sum(1 for e in events if e["teritoriu"] == "da")])
    ws2.append(["Parțial (graniță/zonă învecinată)", sum(1 for e in events if e["teritoriu"] == "partial")])
    ws2.append(["Context regional (în afara granițelor)", sum(1 for e in events if e["teritoriu"] == "nu")])
    ws2.append([])
    ws2.append(["Evenimente pe secole", "Număr"])
    by_century = sorted(Counter(e["secol"] for e in events).items(), key=lambda kv: kv[1], reverse=True)
    for s, n in by_century:
        ws2.append([s, n])
    ws2.append([])
    ws2.append(["Evenimente pe tipuri", "Număr"])
    by_type = sorted(Counter(e["tip"] for e in events).items(), key=lambda kv: kv[1], reverse=True)
    for t, n in by_type:
        ws2.append([t, n])
    ws2.append([])
    ws2.append(["Evenimente pe regiuni", "Număr"])
    by_region = sorted(Counter(e["regiune"] for e in events).items(), key=lambda kv: kv[1], reverse=True)
    for r, n in by_region:
        ws2.append([r, n])
    for c in ws2[1]:
        c.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 60
    ws2.column_dimensions["B"].width = 16

    # ---- Foaia 3: Dicționar câmpuri
    ws3 = wb.create_sheet("Dicționar")
    ws3.append(["Câmp", "Coloană Excel", "Descriere"])
    for k, h, d in FIELDS:
        ws3.append([k, h, d])
    ws3.append(["", "teritoriu = Parțial", "Eveniment la graniță / cu teatru în imediata vecinătate"])
    ws3.append(["", "teritoriu = Nu", "Eveniment de context, cu participare românească, dar în afara granițelor actuale"])
    for c in ws3[1]:
        c.font = Font(bold=True)
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["C"].width = 90

    path = os.path.join(OUT_DIR, "conflicte_militare_romania.xlsx")
    wb.save(path)
    return path


def write_report(events):
    total = len(events)
    on_territory = sum(1 for e in events if e["teritoriu"] == "da")
    partial = sum(1 for e in events if e["teritoriu"] == "partial")
    context = sum(1 for e in events if e["teritoriu"] == "nu")
    approx_locations = sum(1 for e in events if e["zona_aprox"])
    with_coords = sum(1 for e in events if e.get("lat") is not None and e.get("lng") is not None)

    lines = ["""# Baza de date a conflictelor militare de pe teritoriul României
**Sec. VIII î.Hr. – 1944** · generată din `batalii.csv` (28 de secole, de la 8 î.Hr. la 20 d.Hr.)

## Fișiere generate
| Fișier | Descriere |
|---|---|
| `conflicte_militare_romania.csv` | Tabelul principal (UTF-8 cu BOM, deschide corect în Excel/Sheets) |
| `conflicte_militare_romania.json` | Aceleași date în JSON (pentru aplicații web) |
| `conflicte_militare_romania.xlsx` | Excel cu 3 foi: Conflicte, Statistici, Dicționar |
| `RAPORT_STRUCTURARE.md` | Acest raport |

## Statistici
| Indicator | Valoare |
|---|---|
| Total evenimente | **{total}** |
| Pe teritoriul actual al României | {on_territory} |
| Parțial (graniță / zona învecinată) | {partial} |
| Context regional (participare românească, în afara granițelor) | {context} |
| Evenimente cu locație aproximativă (zona) | {approx_locations} |
| Evenimente cu coordonate | {with_coords} |

## Structura unei înregistrări
| Câmp | Descriere |
|---|---|
| `titlu` | Numele bătăliei / conflictului |
| `tip` | Tipologia (bătălie, invazie, asediu, revoltă, bombardament etc.) |
| `data_start` / `data_end` | Datare umană; `c.` = circa |
| `an_start` / `an_end` | Numeric (negativ = î.Hr.), pentru sortare/filtrare |
| `secol` / `secol_n` | Secolul text și numeric |
| `locatie` | Localizare, iar unde nu se știe exact — zona aproximativă |
| `zona_aprox` | 0 = localizare relativ precisă, 1 = zonă aproximativă/incertă |
| `judet` / `regiune` | Județul actual și regiunea istorică |
| `lat` / `lng` | Coordonate aproximative WGS84 (pentru hartă) |
| `participanti` | Beligeranți normalizați în română |
| `descriere` | Context + desfășurare |
| `rezultat` | Cine a câștigat / consecințe |
| `teritoriu` | `da` / `partial` / `nu` (vezi mai jos) |
| `observatii` | Incertitudini de datare/locație |
| `sursa` | Secțiunea din `batalii.csv` de proveniență |

## Note metodologice
1. **Sursă unică de plecare**: coloana `REZULTAT` din `batalii.csv`, care conține, pentru fiecare secol, narativele conflictelor. Evenimentele au fost desfăcute în înregistrări individuale (de exemplu, „Războaiele Daco-Romane” → război 101–102, Tapae 101, Adamclisi, război 105–106, căderea Sarmizegetusei).
2. **Datare**: am păstrat datarea din sursă; unde era expresă, am păstrat ziua/luna (ex. 9–12 noiembrie 1330 pentru Posada). Unde sursa folosește „cca./aprox.”, am marcat cu `c.` și `zona_aprox`.
3. **Locație**: am adăugat județul actual și coordonate aproximative (WGS84) acolo unde toponimul este identificabil. Pentru evenimente fără localizare precisă (ex. Nedao, Posada, Rovine), locația este descrisă ca zonă aproximativă și marcată cu `zona_aprox=1`, iar disputele sunt consemnate în `observatii`.
4. **`teritoriu = nu`**: evenimente de context cu participare românească, dar desfășurate în afara granițelor actuale (ex. Naissus 269, Nicopole 1396, Mohács 1526, Hotin 1621, Turtucaia 1916). Nu sunt incluse în „baza de conflicte de pe teritoriu”, dar sunt utile pentru context.
5. **Diacritice**: fișierul `batalii.csv` a fost livrat în encoding cp1252 cu diacriticele pierdute (literele ă/â/î/ș/ț deveniseră `?`). Textul a fost re-romanizat pe baza contextului istoric și a toponimelor cunoscute; greșelile reziduale de diacritice trebuie verificate față de sursa originală.
6. **Surse secundare**: pentru completarea locațiilor, județelor și a diacriticelor am folosit cunoștințe istorice generale și toponimia standard; nu a fost folosită internetul în acest pas. Evenimentele cu datare/locație disputată sunt marcate în `observatii`.

## Evenimente cu datare sau locație disputată (de verificat)
""".format(total=total, on_territory=on_territory, partial=partial, context=context,
             approx_locations=approx_locations, with_coords=with_coords)]
    disputed = [e for e in events if e.get("observatii") and e["observatii"] != "—"]
    for e in disputed:
        lines.append(f"- **{e['id']}** {e['titlu']} · {e['data_start']} — {e['observatii']}")
    lines.append("""
## Cum poate fi folosită
- **Excel**: filtrați după `Secol`, `Tip conflict` sau `Regiune istorică`.
- **Hartă**: folosiți coloanele `lat`/`lng` (unele coordonate sunt aproximative pentru evenimente preistorice/nomade — verificați `zona_aprox`).
- **Cronologie**: sortați după `an_start`/`an_end` (negativ = î.Hr.).
- **Aplicația DetectLab**: conținutul din `conflicte_militare_romania.json` poate fi încărcat ca dataset separat; nu trebuie importat în tabelul `public.events`, al cărui schema este destinat evenimentelor comunitare (creator_id, pin_id, etc.).
""".format(total=total))
    path = os.path.join(OUT_DIR, "RAPORT_STRUCTURARE.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


def main():
    events = build_events()
    csv_path = write_csv(events)
    json_path = write_json(events)
    xlsx_path = write_xlsx(events)
    report_path = write_report(events)
    print("Evenimente:", len(events))
    print("CSV   :", csv_path)
    print("JSON  :", json_path)
    print("XLSX  :", xlsx_path)
    print("Raport:", report_path)
    # câteva statistici utile în consolă
    print("\nPe secole:")
    for s, n in sorted(Counter(e["secol"] for e in events).items()):
        print(f"  {s}: {n}")
    print("\nPe tipuri (top 12):")
    for t, n in Counter(e["tip"] for e in events).most_common(12):
        print(f"  {t}: {n}")
    print("\nPe teritoriu:", Counter(e["teritoriu"] for e in events))
    print("Locații aproximative:", sum(1 for e in events if e["zona_aprox"]))


if __name__ == "__main__":
    main()
